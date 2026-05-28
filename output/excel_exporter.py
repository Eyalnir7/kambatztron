"""Excel exporter for schedules.

Writes one worksheet per job type, with job names as rows, time slots as
columns, and assigned cadet names colored by team/platoon combination.
"""

from collections import defaultdict
from dataclasses import dataclass
from hashlib import sha1
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from domain.cadet import Cadet
from domain.job import Shift
from scheduling.schedule import Schedule


@dataclass(frozen=True)
class _ColorStyle:
    fill: PatternFill
    font: Font


class ExcelExporter:
    """Export a schedule to an `.xlsx` workbook.

    If the output path ends in `.xlsx`, a workbook is written there. If the
    path has another suffix or no suffix, a CSV fallback directory is created
    for backwards compatibility.
    """

    _PALETTE = [
        "FDE68A",
        "BFDBFE",
        "BBF7D0",
        "FECACA",
        "E9D5FF",
        "FED7AA",
        "A7F3D0",
        "FBCFE8",
        "C7D2FE",
        "F9A8D4",
        "FCD34D",
        "86EFAC",
        "93C5FD",
        "FCA5A5",
        "DDD6FE",
    ]

    def export(self, schedule: Schedule, cadets: List[Cadet], output_path: str) -> None:
        output = Path(output_path)
        if output.suffix.lower() == ".xlsx":
            self._export_workbook(schedule, cadets, output)
            return

        self._export_csv_fallback(schedule, output)

    def _export_workbook(self, schedule: Schedule, cadets: List[Cadet], output: Path) -> None:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        cadet_lookup = {cadet.personal_number: cadet for cadet in cadets}
        color_map = self._build_color_map(cadets)

        grouped_shifts: Dict[str, List[Shift]] = defaultdict(list)
        for shift in schedule.assignments:
            grouped_shifts[shift.job.job_type].append(shift)

        for job_type in sorted(grouped_shifts):
            self._write_job_type_sheet(
                workbook=workbook,
                sheet_name=job_type,
                shifts=grouped_shifts[job_type],
                cadet_lookup=cadet_lookup,
                color_map=color_map,
            )

        self._write_legend_sheet(workbook, cadets, color_map)
        output.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output)

    def _export_csv_fallback(self, schedule: Schedule, output: Path) -> None:
        """Keep the legacy CSV directory behavior for non-XLSX outputs."""
        import csv

        output_dir = output.with_suffix("") if output.suffix else output
        output_dir.mkdir(parents=True, exist_ok=True)

        grouped_shifts: Dict[str, List[Shift]] = defaultdict(list)
        for shift in schedule.assignments:
            grouped_shifts[shift.job.job_type].append(shift)

        for job_type, shifts in grouped_shifts.items():
            rows = self._build_table_rows(shifts)
            file_path = output_dir / f"{self._sanitize_filename(job_type)}.csv"
            with open(file_path, "w", encoding="utf-8", newline="") as file_obj:
                writer = csv.writer(file_obj)
                writer.writerow(["job_name"] + rows["columns"])
                for job_name in rows["row_order"]:
                    row_values = [rows["cells"].get(job_name, {}).get(column, "") for column in rows["columns"]]
                    writer.writerow([job_name] + row_values)

    def _write_job_type_sheet(
        self,
        workbook: Workbook,
        sheet_name: str,
        shifts: List[Shift],
        cadet_lookup: Dict[str, Cadet],
        color_map: Dict[Tuple[str, str], _ColorStyle],
    ) -> None:
        sheet = workbook.create_sheet(title=self._unique_sheet_name(workbook, sheet_name))
        rows = self._build_table_rows(shifts)

        sheet.freeze_panes = "B2"
        sheet.cell(row=1, column=1, value="job_name").font = Font(bold=True)
        sheet.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

        for column_index, column_value in enumerate(rows["columns"], start=2):
            cell = sheet.cell(row=1, column=column_index, value=column_value)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_index, job_name in enumerate(rows["row_order"], start=2):
            label_cell = sheet.cell(row=row_index, column=1, value=job_name)
            label_cell.font = Font(bold=True)
            label_cell.alignment = Alignment(horizontal="left", vertical="center")

            for column_index, column_value in enumerate(rows["columns"], start=2):
                assigned_name = rows["cells"].get(job_name, {}).get(column_value, "")
                cell = sheet.cell(row=row_index, column=column_index, value=assigned_name)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if not assigned_name:
                    continue

                cadet = self._find_cadet_by_name(cadet_lookup.values(), assigned_name)
                if cadet is None:
                    continue

                style = color_map[(cadet.team, cadet.platoon)]
                cell.fill = style.fill
                cell.font = style.font

        self._autosize_sheet(sheet)

    def _write_legend_sheet(
        self,
        workbook: Workbook,
        cadets: List[Cadet],
        color_map: Dict[Tuple[str, str], _ColorStyle],
    ) -> None:
        sheet = workbook.create_sheet(title=self._unique_sheet_name(workbook, "Legend"))
        sheet.freeze_panes = "A2"

        headers = ["team", "platoon", "cadets", "color"]
        for column_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=column_index, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        combos: Dict[Tuple[str, str], List[str]] = defaultdict(list)
        for cadet in cadets:
            combos[(cadet.team, cadet.platoon)].append(cadet.name)

        for row_index, ((team, platoon), cadet_names) in enumerate(sorted(combos.items()), start=2):
            style = color_map[(team, platoon)]
            sheet.cell(row=row_index, column=1, value=team)
            sheet.cell(row=row_index, column=2, value=platoon)
            sheet.cell(row=row_index, column=3, value=", ".join(sorted(cadet_names)))
            color_cell = sheet.cell(row=row_index, column=4, value=style.fill.fgColor.rgb[-6:])
            color_cell.fill = style.fill
            color_cell.font = style.font

        self._autosize_sheet(sheet)

    def _build_color_map(self, cadets: Iterable[Cadet]) -> Dict[Tuple[str, str], _ColorStyle]:
        color_map: Dict[Tuple[str, str], _ColorStyle] = {}
        seen: List[Tuple[str, str]] = []

        for cadet in cadets:
            key = (cadet.team, cadet.platoon)
            if key in color_map:
                continue

            palette_index = self._palette_index(key, len(seen))
            hex_color = self._PALETTE[palette_index]
            font_color = self._contrasting_font_color(hex_color)
            color_map[key] = _ColorStyle(
                fill=PatternFill(fill_type="solid", fgColor=hex_color),
                font=Font(color=font_color),
            )
            seen.append(key)

        return color_map

    def _build_table_rows(self, shifts: List[Shift]) -> Dict[str, object]:
        columns: List[str] = []
        row_order: List[str] = []
        cells: Dict[str, Dict[str, str]] = defaultdict(dict)

        for shift in shifts:
            time_slot = str(shift.time_slot)
            if time_slot not in columns:
                columns.append(time_slot)
            if shift.job.name not in row_order:
                row_order.append(shift.job.name)
            cells[shift.job.name][time_slot] = shift.assigned_cadet.name if shift.assigned_cadet else ""

        columns = sorted(columns, key=self._time_slot_sort_key)
        return {"columns": columns, "row_order": row_order, "cells": cells}

    def _find_cadet_by_name(self, cadets: Iterable[Cadet], name: str) -> Optional[Cadet]:
        for cadet in cadets:
            if cadet.name == name:
                return cadet
        return None

    def _palette_index(self, key: Tuple[str, str], salt: int) -> int:
        digest = sha1(f"{key[0]}::{key[1]}::{salt}".encode("utf-8")).digest()
        return digest[0] % len(self._PALETTE)

    def _contrasting_font_color(self, hex_color: str) -> str:
        red = int(hex_color[0:2], 16)
        green = int(hex_color[2:4], 16)
        blue = int(hex_color[4:6], 16)
        luminance = (0.299 * red) + (0.587 * green) + (0.114 * blue)
        return "000000" if luminance > 170 else "FFFFFF"

    def _time_slot_sort_key(self, value: str) -> Tuple[datetime, datetime]:
        start_text = value[:16]
        end_text = value[17:33]
        return (
            datetime.strptime(start_text, "%Y-%m-%d %H:%M"),
            datetime.strptime(end_text, "%Y-%m-%d %H:%M"),
        )

    def _autosize_sheet(self, sheet) -> None:
        for column_cells in sheet.columns:
            column_letter = column_cells[0].column_letter
            max_length = 0
            for cell in column_cells:
                if cell.value is None:
                    continue
                max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 40)

    def _unique_sheet_name(self, workbook: Workbook, desired_name: str) -> str:
        base_name = self._sanitize_sheet_name(desired_name)
        candidate = base_name
        suffix = 1
        while candidate in workbook.sheetnames:
            truncated = base_name[: max(0, 31 - len(f"_{suffix}"))]
            candidate = f"{truncated}_{suffix}"
            suffix += 1
        return candidate

    def _sanitize_sheet_name(self, value: str) -> str:
        invalid = ["/", "\\", "?", "*", "[", "]", ":"]
        cleaned = value
        for character in invalid:
            cleaned = cleaned.replace(character, "-")
        return cleaned[:31] if len(cleaned) > 31 else cleaned

    def _sanitize_filename(self, value: str) -> str:
        return self._sanitize_sheet_name(value).replace(" ", "_")
